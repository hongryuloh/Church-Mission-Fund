import streamlit as st
import pandas as pd
import io
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.worksheet.pagebreak import Break
from openpyxl.utils import get_column_letter
from datetime import datetime
from sqlalchemy import text

# --- 1. 앱 기본 설정 및 CSS ---
st.set_page_config(page_title="선교헌금 관리", layout="wide", initial_sidebar_state="auto")

st.markdown("""
    <style>
    .block-container { padding-top: 2rem !important; padding-bottom: 5rem !important; padding-left: 1rem !important; padding-right: 1rem !important; }
    #MainMenu { visibility: hidden; display: none !important; }
    header { background-color: rgba(0,0,0,0) !important; }
    h1 { font-size: 1.7rem !important; white-space: nowrap !important; text-align: center !important; margin-bottom: 1rem !important; letter-spacing: -1px; }
    div[data-testid="stSidebarCollapsedControl"] button { background-color: #ff4b4b !important; color: white !important; border-radius: 50% !important; width: 48px !important; height: 48px !important; position: fixed !important; top: 15px !important; left: 15px !important; z-index: 999999 !important; box-shadow: 0 4px 10px rgba(0,0,0,0.3) !important; border: 2px solid white !important; }
    div[data-testid="stSidebarCollapsedControl"] button svg { fill: white !important; width: 26px !important; height: 26px !important; }
    div[data-testid="stHorizontalBlock"] { display: flex !important; flex-direction: row !important; flex-wrap: nowrap !important; align-items: flex-end !important; gap: 5px !important; }
    div[data-testid="stHorizontalBlock"] > div { flex: 1 1 auto !important; min-width: 0 !important; }
    .fixed-footer { position: fixed; bottom: 0; left: 0; width: 100%; background-color: white; padding: 10px 15px 30px 15px; border-top: 1px solid #ddd; z-index: 999; }
    @media (prefers-color-scheme: dark) { .fixed-footer { background-color: #1e1e1e !important; border-top: 1px solid #333 !important; } }
    .stButton button { width: 100% !important; padding: 0px !important; font-size: 13px !important; height: 42px !important; }
    .stNumberInput input { height: 42px !important; }
    [data-testid="stSidebar"] .stRadio div[role="radiogroup"] label { font-size: 16px !important; font-weight: 600 !important; padding: 10px 0px !important; }
    </style>
    """, unsafe_allow_html=True)

if "authenticated" not in st.session_state: st.session_state["authenticated"] = False
if "current_user" not in st.session_state: st.session_state["current_user"] = ""

# --- 2. DB 연결 및 초고속 캐싱 시스템 적용 ---
conn = st.connection("postgresql", type="sql")

# DB에서 타겟 연도 가져오기 (메모리에 캐싱하여 속도 극대화)
@st.cache_data(ttl=600)
def get_target_year():
    try:
        res = conn.query("SELECT setting_value FROM settings WHERE setting_key='TARGET_YEAR'", ttl=0)
        if not res.empty: return int(res.iloc[0]['setting_value'])
    except: pass
    return 2026

TARGET_YEAR = get_target_year()

# 로그인 화면
if not st.session_state["authenticated"]:
    col1, col2, col3 = st.columns([0.1, 3, 0.1]) 
    with col2:
        st.title(f"⛪ {TARGET_YEAR} 선교헌금 관리")
        st.info("🔒 ID와 비밀번호를 입력해 주세요.")
        with st.form("login_form"):
            input_id = st.text_input("아이디 (ID)")
            input_pwd = st.text_input("비밀번호 (Password)", type="password")
            if st.form_submit_button("로그인", use_container_width=True):
                credentials = st.secrets.get("credentials", {})
                if input_id in credentials and str(credentials[input_id]) == input_pwd:
                    st.session_state["authenticated"] = True; st.session_state["current_user"] = input_id; st.rerun() 
                else: st.error("❌ 정보가 올바르지 않습니다.")
    st.stop() 

with st.sidebar:
    st.title(f"⛪ {TARGET_YEAR} 선교헌금")
    st.write(f"👤 **{st.session_state['current_user']}**님")
    if st.button("로그아웃", use_container_width=True):
        st.session_state["authenticated"] = False; st.session_state["current_user"] = ""; st.rerun()
    st.markdown("---") 

# 🚀 데이터를 메모리에 저장해두고 꺼내쓰는 초고속 로딩 함수
@st.cache_data(ttl=600)
def load_data(year):
    df_inc = conn.query("SELECT * FROM income ORDER BY date DESC, id DESC;", ttl=0)
    df_inc = df_inc.rename(columns={'date': '날짜', 'year_month': '년월', 'name': '이름', 'amount': '금액', 'note': '비고'})
    
    df_exp = conn.query("SELECT * FROM expense ORDER BY date DESC, id DESC;", ttl=0)
    df_exp = df_exp.rename(columns={'date': '날짜', 'year_month': '년월', 'item': '내역', 'amount': '금액', 'note': '비고'})
    
    # 💡 작정액은 설정된 TARGET_YEAR에 해당하는 성도만 가져옴
    df_tgt = conn.query(f"SELECT * FROM target WHERE target_year={year} ORDER BY name ASC;", ttl=0)
    df_tgt = df_tgt.rename(columns={'name': '이름', 'role': '직분', 'monthly_amount': '월별 작정액', 'print_yn': '인쇄여부'})
    
    df_cat = conn.query(f"SELECT * FROM expense_category WHERE target_year={year} ORDER BY item_name ASC;", ttl=0)
    
    return df_inc, df_tgt, df_exp, df_cat

# DB에 데이터가 변경될 때 캐시를 초기화하는 함수
def clear_db_cache(): st.cache_data.clear()

def format_date_str(x):
    if pd.isna(x): return ""
    if isinstance(x, pd.Timestamp) or hasattr(x, 'strftime'): return x.strftime('%Y-%m-%d')
    return str(x)

def fmt(val): return "-" if pd.isna(val) or val == 0 else f"{int(val):,}"

# --- 3. 데이터 계산 함수 ---
def calculate_details(user_name, df_income, df_target, start_year=TARGET_YEAR):
    user_info = df_target[df_target['이름'].astype(str).str.strip() == user_name.strip()]
    if user_info.empty: return None
    commit = float(user_info.iloc[0].get('월별 작정액', 0))
    total_donated = float(user_info.iloc[0].get('헌금액', 0))
    
    u_inc = df_income[df_income['이름'].astype(str).str.strip() == user_name.strip()].copy()
    u_inc['금액'] = pd.to_numeric(u_inc['금액'], errors='coerce').fillna(0)
    paid = u_inc.groupby('년월')['금액'].sum().to_dict()
    
    alloc, lab = [0.0]*13, [""]*13
    sorted_m = sorted([m for m in paid.keys() if str(m).startswith(str(start_year))])
    
    if commit > 0:
        ptr = 1
        for pm in sorted_m:
            val = float(paid.get(pm, 0))
            while val > 0.01 and ptr <= 12:
                rem = commit - alloc[ptr]
                if rem <= 0.01: ptr += 1
                else:
                    amt = min(val, rem); txt = f"{int(str(pm)[-2:]):02d}월납"
                    lab[ptr] = txt if lab[ptr] == "" else f"{lab[ptr]}\n{txt}"
                    alloc[ptr] += amt; val -= amt
                    if alloc[ptr] >= commit - 0.01: ptr += 1
    else:
        for pm in sorted_m:
            try:
                m = int(str(pm)[-2:])
                if 1 <= m <= 12: alloc[m] = float(paid.get(pm, 0)); lab[m] = f"{m:02d}월납"
            except: pass 
    return {"name": user_name, "commit": commit, "alloc": alloc[1:], "labs": lab[1:], "total": total_donated}

# --- 4. 인쇄 포맷 ---
def generate_summary_excel(df_income, df_target, target_month, start_year=TARGET_YEAR):
    wb = openpyxl.Workbook(); ws = wb.active; ws.title = "개인별 헌금내역"
    for c in range(1, 19): ws.column_dimensions[get_column_letter(c)].width = 8.13
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    print_users = df_target[df_target['인쇄여부'] == 'Y'] 
    user_list = []
    for _, row_data in print_users.iterrows():
        name, pos = str(row_data.get('이름', '')).strip(), str(row_data.get('직분', '')).strip()
        if not name or name == 'nan' or name == '합계': continue
        res = calculate_details(name, df_income, df_target, start_year)
        if res: res['pos'] = pos; user_list.append(res)
        
    today_str = datetime.now().strftime("%Y.%m.%d")
    current_row = 1
    
    def draw_user_block(r, c_off, user):
        for r_i in range(r, r+9): ws.row_dimensions[r_i].height = 25 
        ws.merge_cells(start_row=r, start_column=1+c_off, end_row=r, end_column=8+c_off)
        ws.cell(row=r, column=1+c_off, value=f"{TARGET_YEAR}년 선교헌금 작정 및 헌금내역").font = Font(size=16, bold=True, underline="single")
        ws.cell(row=r, column=1+c_off).alignment = center_align
        ws.merge_cells(start_row=r+1, start_column=1+c_off, end_row=r+1, end_column=3+c_off); ws.cell(row=r+1, column=1+c_off, value=f"({today_str} 기준)").font = Font(bold=True)
        ws.merge_cells(start_row=r+1, start_column=4+c_off, end_row=r+1, end_column=6+c_off); ws.cell(row=r+1, column=4+c_off, value="선교헌금 합계 :").alignment = Alignment(horizontal='right', vertical='center')
        ws.merge_cells(start_row=r+1, start_column=7+c_off, end_row=r+1, end_column=8+c_off); t_c = ws.cell(row=r+1, column=7+c_off, value=user['total']); t_c.font, t_c.number_format, t_c.alignment = Font(bold=True), '#,##0', center_align
        for row_i in range(r+2, r+8):
            for col_i in range(1+c_off, 9+c_off): ws.cell(row=row_i, column=col_i).border = thin_border; ws.cell(row=row_i, column=col_i).alignment = center_align
        ws.cell(row=r+2, column=1+c_off, value="이름").font = Font(bold=True); ws.cell(row=r+2, column=2+c_off, value="월작정액").font = Font(bold=True)
        for i, m in enumerate(["01", "02", "03", "04", "05", "06"], 3): ws.cell(row=r+2, column=i+c_off, value=m).font = Font(bold=True)
        for i, m in enumerate(["07", "08", "09", "10", "11", "12"], 3): ws.cell(row=r+5, column=i+c_off, value=m).font = Font(bold=True)
        ws.merge_cells(start_row=r+3, start_column=1+c_off, end_row=r+7, end_column=1+c_off); ws.cell(row=r+3, column=1+c_off, value=f"{user['name']}\n{user['pos']}").font = Font(bold=True)
        ws.merge_cells(start_row=r+3, start_column=2+c_off, end_row=r+7, end_column=2+c_off); cc = ws.cell(row=r+3, column=2+c_off, value=user['commit'] if user['commit'] > 0 else "-")
        if user['commit'] > 0: cc.number_format = '#,##0'
        for i in range(6):
            ws.cell(row=r+3, column=i+3+c_off, value=user['labs'][i]); c1 = ws.cell(row=r+4, column=i+3+c_off, value=user['alloc'][i] if user['alloc'][i] > 0 else "")
            if user['alloc'][i] > 0: c1.number_format = '#,##0'
            ws.cell(row=r+6, column=i+3+c_off, value=user['labs'][i+6]); c2 = ws.cell(row=r+7, column=i+3+c_off, value=user['alloc'][i+6] if user['alloc'][i+6] > 0 else "")
            if user['alloc'][i+6] > 0: c2.number_format = '#,##0'
        ws.merge_cells(start_row=r+8, start_column=1+c_off, end_row=r+8, end_column=8+c_off); ws.cell(row=r+8, column=1+c_off, value="선교헌금에 관심가져주셔서 감사합니다.").alignment = center_align

    for i in range(0, len(user_list), 2):
        user_left, user_right = user_list[i], user_list[i+1] if i+1 < len(user_list) else None
        for r_idx in range(current_row, current_row + 9):
            ws.row_dimensions[r_idx].height = 25; ws.cell(row=r_idx, column=9).border = Border(right=Side(style='thin', color='000000'))
        draw_user_block(current_row, 0, user_left) 
        if user_right: draw_user_block(current_row, 10, user_right)
        pair_index = i // 2
        if pair_index % 2 == 0:
            gap_start = current_row + 9
            for r_idx in range(gap_start, gap_start + 3): ws.row_dimensions[r_idx].height = 25; ws.cell(row=r_idx, column=9).border = Border(right=Side(style='thin', color='000000'))
            for col_i in range(1, 19): ws.cell(row=gap_start, column=col_i).border = Border(bottom=Side(style='dashed', color='888888'), right=ws.cell(row=gap_start, column=col_i).border.right)
            current_row += 12 
        else:
            gap_start = current_row + 9; ws.row_dimensions[gap_start].height = 25; ws.cell(row=gap_start, column=9).border = Border(right=Side(style='thin', color='000000'))
            for col_i in range(1, 19): ws.cell(row=gap_start, column=col_i).border = Border(bottom=Side(style='dashed', color='888888'), right=ws.cell(row=gap_start, column=col_i).border.right)
            ws.row_breaks.append(Break(id=gap_start)); current_row += 10 
    ws.page_setup.orientation, ws.page_setup.fitToPage, ws.page_setup.fitToWidth = ws.ORIENTATION_LANDSCAPE, True, 1  
    ws.page_margins.left = ws.page_margins.right = 0.25
    output = io.BytesIO(); wb.save(output); return output.getvalue()

# --- 5. 앱 화면 구성 ---
for k in ['edit_idx_inc', 'edit_idx_exp', 'edit_idx_tgt', 'mode_inc', 'mode_exp', 'mode_tgt']:
    if k not in st.session_state: st.session_state[k] = None

df_income, df_target, df_expense, df_cat = load_data(TARGET_YEAR)

# 메뉴 권한 설정 (관리자는 '시스템 설정' 메뉴 추가)
user_id = st.session_state["current_user"]
if user_id == "admin":
    menu_options = ["✍️ 데이터 관리", "📊 결산/주단위집계", "🔍 개인별 조회", "🖨️ 인쇄용 집계표", "⚙️ 시스템 설정"]
elif user_id == "mission01":
    menu_options = ["✍️ 데이터 관리", "📊 결산/주단위집계", "🔍 개인별 조회"]
else:
    menu_options = ["📊 결산/주단위집계"]
menu = st.sidebar.radio("메뉴 선택", menu_options)

# 0. 시스템 설정 (새 기능: 연도 및 지출항목 관리)
if menu == "⚙️ 시스템 설정":
    st.header("⚙️ 시스템 환경 및 항목 관리")
    st.info("💡 여기서 연도를 변경하면 해당 연도의 작정액과 지출항목 데이터를 자동으로 불러옵니다.")
    
    tab_y, tab_c = st.tabs(["🗓️ 연도 설정", "📋 지출항목 관리"])
    
    with tab_y:
        st.subheader("기준 연도 설정 (TARGET YEAR)")
        with st.form("year_setting_form"):
            new_year = st.number_input("목표 연도 입력", value=TARGET_YEAR, step=1)
            if st.form_submit_button("연도 변경 저장", use_container_width=True):
                with conn.session as s:
                    s.execute(text("UPDATE settings SET setting_value=:y WHERE setting_key='TARGET_YEAR'"), {"y": str(new_year)})
                    s.commit()
                clear_db_cache(); st.success(f"✅ {new_year}년으로 변경되었습니다!"); st.rerun()

    with tab_c:
        st.subheader(f"[{TARGET_YEAR}년] 지출 항목 관리")
        df_cat_view = df_cat[['id', 'item_name']].rename(columns={'id': 'ID', 'item_name': '지출항목명'})
        st.dataframe(df_cat_view, use_container_width=True, hide_index=True)
        
        c1, c2 = st.columns(2)
        with c1:
            with st.form("cat_add_form"):
                new_cat = st.text_input("➕ 새 지출항목 추가")
                if st.form_submit_button("항목 저장", use_container_width=True):
                    if new_cat:
                        # 💡 1. 이미 표에 있는 이름인지 먼저 검사합니다.
                        if new_cat.strip() in df_cat['item_name'].tolist():
                            st.warning("⚠️ 이미 목록에 존재하는 항목입니다.")
                        else:
                            # 💡 2. 목록에 없으면 DB에 안전하게 저장하고 새로고침!
                            with conn.session as s:
                                s.execute(text("INSERT INTO expense_category (target_year, item_name) VALUES (:y, :n)"), {"y": TARGET_YEAR, "n": new_cat.strip()})
                                s.commit()
                            clear_db_cache()
                            st.rerun()
        with c2:
            with st.form("cat_del_form"):
                del_id = st.number_input("🗑️ 삭제할 항목 ID", min_value=0, step=1)
                if st.form_submit_button("항목 삭제", use_container_width=True):
                    with conn.session as s:
                        s.execute(text("DELETE FROM expense_category WHERE id=:id"), {"id": del_id})
                        s.commit()
                    clear_db_cache()
                    st.rerun()

# 1. 데이터 관리
elif menu == "✍️ 데이터 관리":
    tab1, tab2, tab3 = st.tabs(["💰 헌금 수입", "📉 지출 내역", "👤 작정액 관리"])
    
    with tab1: 
        if st.session_state.mode_inc is None:
            df_view = df_income.copy()
            if '금액' in df_view.columns: df_view['금액'] = pd.to_numeric(df_view['금액'], errors='coerce').fillna(0).apply(lambda x: f"{int(x):,} 원")
            st.dataframe(df_view[['id', '날짜', '이름', '금액', '비고']], use_container_width=True, height=330, hide_index=True)
            
            with st.container():
                st.markdown('<div class="fixed-footer">', unsafe_allow_html=True)
                bc1, bc2, bc3, bc4 = st.columns([1.2, 1, 1, 1])
                with bc1: 
                    if st.button("➕신규", key="inc_new", use_container_width=True): st.session_state.mode_inc = 'add'; st.rerun()
                with bc2: 
                    idx = st.number_input("ID", min_value=0, step=1, key="inc_idx_in", label_visibility="collapsed")
                with bc3: 
                    if st.button("📝수정", key="inc_edit", use_container_width=True): 
                        if not df_income[df_income['id'] == idx].empty: st.session_state.edit_idx_inc, st.session_state.mode_inc = idx, 'edit'; st.rerun()
                        else: st.error("ID 없음")
                with bc4: 
                    if st.button("🗑️삭제", key="inc_del", use_container_width=True): 
                        if not df_income[df_income['id'] == idx].empty: st.session_state.edit_idx_inc, st.session_state.mode_inc = idx, 'delete_check'; st.rerun()
                        else: st.error("ID 없음")
                st.markdown('</div>', unsafe_allow_html=True)
                
        elif st.session_state.mode_inc == 'add':
            with st.form("inc_add"):
                d = st.date_input("입금일자")
                opts = [f"{r['이름']} ({r['직분']})" if pd.notna(r.get('직분')) else str(r.get('이름')) for _, r in df_target.iterrows() if pd.notna(r.get('이름'))]
                sel = st.selectbox("이름 선택", opts) if opts else st.selectbox("이름 선택", ["등록된 성도가 없습니다"])
                amt = st.number_input("금액", min_value=0, step=10000)
                note = st.text_input("비고")
                if st.form_submit_button("저장"):
                    if opts:
                        with conn.session as s:
                            s.execute(text("INSERT INTO income (date, year_month, name, amount, note) VALUES (:d, :ym, :n, :a, :nt)"),
                                      {"d": d, "ym": d.strftime("%Y%m"), "n": sel.split(" (")[0], "a": amt, "nt": note})
                            s.commit()
                        clear_db_cache(); st.session_state.mode_inc = None; st.rerun()
                if st.form_submit_button("취소"): st.session_state.mode_inc = None; st.rerun()
                
        elif st.session_state.mode_inc == 'edit':
            curr = df_income[df_income['id'] == st.session_state.edit_idx_inc].iloc[0]
            with st.form("inc_edit"):
                new_d = st.date_input("날짜", value=pd.to_datetime(curr.get('날짜', datetime.now())))
                new_n = st.text_input("이름", value=str(curr.get('이름', '')))
                new_a = st.number_input("금액", value=int(pd.to_numeric(curr.get('금액', 0), errors='coerce') or 0), step=10000)
                new_b = st.text_input("비고", value=str(curr.get('비고', '')) if pd.notna(curr.get('비고')) else "")
                if st.form_submit_button("✅ 수정 완료"):
                    with conn.session as s:
                        s.execute(text("UPDATE income SET date=:d, year_month=:ym, name=:n, amount=:a, note=:nt WHERE id=:id"),
                                  {"d": new_d, "ym": new_d.strftime("%Y%m"), "n": new_n, "a": new_a, "nt": new_b, "id": st.session_state.edit_idx_inc})
                        s.commit()
                    clear_db_cache(); st.session_state.mode_inc = None; st.rerun()
                if st.form_submit_button("취소"): st.session_state.mode_inc = None; st.rerun()
                
        elif st.session_state.mode_inc == 'delete_check':
            st.warning(f"⚠️ ID {st.session_state.edit_idx_inc} 데이터를 삭제하시겠습니까?")
            if st.button("🔴 삭제 실행", use_container_width=True):
                with conn.session as s:
                    s.execute(text("DELETE FROM income WHERE id=:id"), {"id": st.session_state.edit_idx_inc})
                    s.commit()
                clear_db_cache(); st.session_state.mode_inc = None; st.rerun()
            if st.button("취소", use_container_width=True): st.session_state.mode_inc = None; st.rerun()

    with tab2:
        if st.session_state.mode_exp is None:
            df_exp_v = df_expense.copy()
            if '금액' in df_exp_v.columns: df_exp_v['금액'] = pd.to_numeric(df_exp_v['금액'], errors='coerce').fillna(0).apply(lambda x: f"{int(x):,} 원")
            st.dataframe(df_exp_v[['id', '날짜', '내역', '금액', '비고']], use_container_width=True, height=330, hide_index=True)
            
            with st.container():
                st.markdown('<div class="fixed-footer">', unsafe_allow_html=True)
                ec1, ec2, ec3, ec4 = st.columns([1.2, 1, 1, 1])
                with ec1: 
                    if st.button("➕지출", key="exp_new", use_container_width=True): st.session_state.mode_exp = 'add'; st.rerun()
                with ec2: 
                    idx_e = st.number_input("ID", min_value=0, step=1, key="exp_idx_in", label_visibility="collapsed")
                with ec3: 
                    if st.button("📝수정", key="exp_edit", use_container_width=True): 
                        if not df_expense[df_expense['id'] == idx_e].empty: st.session_state.edit_idx_exp, st.session_state.mode_exp = idx_e, 'edit'; st.rerun()
                with ec4: 
                    if st.button("🗑️삭제", key="exp_del", use_container_width=True): 
                        if not df_expense[df_expense['id'] == idx_e].empty: st.session_state.edit_idx_exp, st.session_state.mode_exp = idx_e, 'delete_check'; st.rerun()
                st.markdown('</div>', unsafe_allow_html=True)

        elif st.session_state.mode_exp == 'add':
            with st.form("exp_add"):
                d = st.date_input("지출일자")
                # 💡 관리자가 지정한 지출항목 리스트만 가져오도록 깔끔하게 변경
                cat_list = df_cat['item_name'].tolist() if not df_cat.empty else ["등록된 항목이 없습니다"]
                sel_item = st.selectbox("지출항목 선택", cat_list)
                amt = st.number_input("금액", min_value=0, step=10000)
                note = st.text_input("비고")
                if st.form_submit_button("저장"):
                    if sel_item == "등록된 항목이 없습니다": st.error("시스템 설정에서 지출항목을 먼저 등록해주세요.")
                    else:
                        with conn.session as s:
                            s.execute(text("INSERT INTO expense (date, year_month, item, amount, note) VALUES (:d, :ym, :i, :a, :nt)"),
                                      {"d": d, "ym": d.strftime("%Y%m"), "i": sel_item, "a": amt, "nt": note})
                            s.commit()
                        clear_db_cache(); st.session_state.mode_exp = None; st.rerun()
                if st.form_submit_button("취소"): st.session_state.mode_exp = None; st.rerun()
                
        elif st.session_state.mode_exp == 'edit':
            curr = df_expense[df_expense['id'] == st.session_state.edit_idx_exp].iloc[0]
            with st.form("exp_edit"):
                new_d = st.date_input("날짜", value=pd.to_datetime(curr.get('날짜', datetime.now())))
                curr_item = str(curr.get('내역', '')).strip()
                cat_list = df_cat['item_name'].tolist() if not df_cat.empty else []
                if curr_item and curr_item not in cat_list: cat_list.append(curr_item)
                default_idx = cat_list.index(curr_item) if curr_item in cat_list else 0
                
                sel_item = st.selectbox("지출항목 선택", cat_list, index=default_idx)
                new_a = st.number_input("금액", value=int(pd.to_numeric(curr.get('금액', 0), errors='coerce') or 0), step=10000)
                new_b = st.text_input("비고", value=str(curr.get('비고', '')) if pd.notna(curr.get('비고')) else "")
                if st.form_submit_button("✅ 수정 완료"):
                    with conn.session as s:
                        s.execute(text("UPDATE expense SET date=:d, year_month=:ym, item=:i, amount=:a, note=:nt WHERE id=:id"),
                                  {"d": new_d, "ym": new_d.strftime("%Y%m"), "i": sel_item, "a": new_a, "nt": new_b, "id": st.session_state.edit_idx_exp})
                        s.commit()
                    clear_db_cache(); st.session_state.mode_exp = None; st.rerun()
                if st.form_submit_button("취소"): st.session_state.mode_exp = None; st.rerun()
                
        elif st.session_state.mode_exp == 'delete_check':
            if st.button("🔴 지출 삭제 실행", use_container_width=True):
                with conn.session as s:
                    s.execute(text("DELETE FROM expense WHERE id=:id"), {"id": st.session_state.edit_idx_exp})
                    s.commit()
                clear_db_cache(); st.session_state.mode_exp = None; st.rerun()
            if st.button("취소", use_container_width=True): st.session_state.mode_exp = None; st.rerun()

    with tab3: 
        if st.session_state.mode_tgt is None:
            df_view = df_target.copy()
            df_view['년간작정금액'] = 0
            df_view['헌금액'] = 0
            df_view['년간작정 잔여금액'] = 0
            
            for idx, row in df_view.iterrows():
                name = str(row.get('이름')).strip()
                if not name or name == 'nan' or name == '합계': continue
                user_inc = df_income[df_income['이름'].apply(lambda x: str(x).strip()) == name]
                total_donated = pd.to_numeric(user_inc['금액'], errors='coerce').sum()
                m_amt = pd.to_numeric(row.get('월별 작정액', 0), errors='coerce') or 0
                df_view.loc[idx, ['년간작정금액', '헌금액', '년간작정 잔여금액']] = [m_amt * 12, total_donated, (m_amt * 12) - total_donated]
            
            for c in ['월별 작정액', '년간작정금액', '헌금액', '년간작정 잔여금액']: 
                if c in df_view.columns: df_view[c] = pd.to_numeric(df_view[c], errors='coerce').fillna(0).apply(lambda x: f"{int(x):,} 원")
            
            st.dataframe(df_view[['id', '이름', '직분', '월별 작정액', '년간작정 잔여금액', '헌금액']], use_container_width=True, height=330, hide_index=True)
            
            with st.container():
                st.markdown('<div class="fixed-footer">', unsafe_allow_html=True)
                tc1, tc2, tc3, tc4 = st.columns([1.2, 1, 1, 1])
                with tc1: 
                    if st.button("➕신규", key="tgt_new", use_container_width=True): st.session_state.mode_tgt = 'add'; st.rerun()
                with tc2: 
                    idx_t = st.number_input("ID", min_value=0, step=1, key="tgt_idx_in", label_visibility="collapsed")
                with tc3: 
                    if st.button("📝수정", key="tgt_edit", use_container_width=True): 
                        if not df_target[df_target['id'] == idx_t].empty: st.session_state.edit_idx_tgt, st.session_state.mode_tgt = idx_t, 'edit'; st.rerun()
                with tc4: 
                    if st.button("🗑️삭제", key="tgt_del", use_container_width=True): 
                        if not df_target[df_target['id'] == idx_t].empty: st.session_state.edit_idx_tgt, st.session_state.mode_tgt = idx_t, 'delete_check'; st.rerun()
                st.markdown('</div>', unsafe_allow_html=True)

        elif st.session_state.mode_tgt == 'add':
            with st.form("tgt_add"):
                roles = ["목사", "사모", "전도사", "장로", "안수집사", "권사", "집사", "성도"]
                n = st.text_input("이름")
                p = st.selectbox("직분", roles, index=7)
                a = st.number_input("월별 작정액", min_value=0, step=10000)
                if st.form_submit_button("저장"):
                    with conn.session as s:
                        # 💡 새 성도를 추가할 때 설정된 TARGET_YEAR 도 함께 저장합니다.
                        s.execute(text("INSERT INTO target (target_year, name, role, monthly_amount, print_yn) VALUES (:y, :n, :r, :a, 'N')"),
                                  {"y": TARGET_YEAR, "n": n, "r": p, "a": a})
                        s.commit()
                    clear_db_cache(); st.session_state.mode_tgt = None; st.rerun()
                if st.form_submit_button("취소"): st.session_state.mode_tgt = None; st.rerun()
                
        elif st.session_state.mode_tgt == 'edit':
            curr = df_target[df_target['id'] == st.session_state.edit_idx_tgt].iloc[0]
            with st.form("tgt_edit"):
                roles = ["목사", "사모", "전도사", "장로", "안수집사", "권사", "집사", "성도"]
                curr_p = str(curr.get('직분', '')).strip()
                if curr_p and curr_p not in roles: roles.append(curr_p)
                default_idx = roles.index(curr_p) if curr_p in roles else 7
                
                n = st.text_input("이름", value=str(curr.get('이름', '')))
                p = st.selectbox("직분", roles, index=default_idx)
                a = st.number_input("월별 작정액", value=int(pd.to_numeric(curr.get('월별 작정액', 0), errors='coerce') or 0), step=10000)
                if st.form_submit_button("✅ 수정 완료"):
                    with conn.session as s:
                        s.execute(text("UPDATE target SET name=:n, role=:r, monthly_amount=:a WHERE id=:id"),
                                  {"n": n, "r": p, "a": a, "id": st.session_state.edit_idx_tgt})
                        s.commit()
                    clear_db_cache(); st.session_state.mode_tgt = None; st.rerun()
                if st.form_submit_button("취소"): st.session_state.mode_tgt = None; st.rerun()
                
        elif st.session_state.mode_tgt == 'delete_check':
            if st.button("🔴 성도 데이터 삭제 확정"):
                with conn.session as s:
                    s.execute(text("DELETE FROM target WHERE id=:id"), {"id": st.session_state.edit_idx_tgt})
                    s.commit()
                clear_db_cache(); st.session_state.mode_tgt = None; st.rerun()

# 2. 결산/주단위집계
elif menu == "📊 결산/주단위집계":
    tab1, tab2 = st.tabs(["📅 월별 결산내역", "📆 주단위 결산내역"])
    df_inc_calc, df_exp_calc = df_income.copy(), df_expense.copy()
    df_inc_calc['amt'] = pd.to_numeric(df_inc_calc['금액'], errors='coerce').fillna(0)
    df_exp_calc['amt'] = pd.to_numeric(df_exp_calc['금액'], errors='coerce').fillna(0)
    
    c_inc = df_inc_calc[(df_inc_calc['날짜'].astype(str) < f"{TARGET_YEAR}-01-01") | (df_inc_calc['이름'].astype(str).str.contains('전년이월'))]['amt'].sum()
    c_exp = df_exp_calc[(df_exp_calc['날짜'].astype(str) < f"{TARGET_YEAR}-01-01") | (df_exp_calc['내역'].astype(str).str.contains('전년이월'))]['amt'].sum()
    carryover_bal = c_inc - c_exp
    
    df_inc_target = df_inc_calc[(df_inc_calc['날짜'].astype(str) >= f"{TARGET_YEAR}-01-01") & (~df_inc_calc['이름'].astype(str).str.contains('전년이월'))]
    df_exp_target = df_exp_calc[(df_exp_calc['날짜'].astype(str) >= f"{TARGET_YEAR}-01-01") & (~df_exp_calc['내역'].astype(str).str.contains('전년이월'))]
    
    with tab1:
        st.subheader(f"{TARGET_YEAR}년 선교헌금 월별결산")
        monthly_data = [{"월별": "전년이월", "수입": carryover_bal, "지출": 0, "잔액": carryover_bal}]
        cur_bal, tot_inc, tot_exp = carryover_bal, carryover_bal, 0
        for m in range(1, 13):
            ym = f"{TARGET_YEAR}{m:02d}"
            inc = df_inc_target[df_inc_target['년월'] == ym]['amt'].sum()
            exp = df_exp_target[df_exp_target['년월'] == ym]['amt'].sum()
            if inc == 0 and exp == 0 and m > datetime.now().month: monthly_data.append({"월별": ym, "수입": 0, "지출": 0, "잔액": 0})
            else: cur_bal += (inc - exp); tot_inc += inc; tot_exp += exp; monthly_data.append({"월별": ym, "수입": inc, "지출": exp, "잔액": cur_bal})
        monthly_data.append({"월별": "합계", "수입": tot_inc, "지출": tot_exp, "잔액": tot_inc - tot_exp})
        
        h1 = "<table style='width:100%; border-collapse: collapse; text-align: center; border: 2px solid #a4b7c6; font-size: 15px; background-color: #ffffff; color: #333333;'>"
        h1 += "<tr style='background-color: #dbe5f1;'><th style='border: 1px solid #a4b7c6; padding: 10px;'>월별</th><th style='border: 1px solid #a4b7c6; padding: 10px;'>수입</th><th style='border: 1px solid #a4b7c6; padding: 10px;'>지출</th><th style='border: 1px solid #a4b7c6; padding: 10px;'>잔액</th></tr>"
        for row in monthly_data:
            bg = "#b4c6e7" if row['월별'] == "합계" else ("#f4f5f7" if row['월별'] == "전년이월" else "#ffffff")
            h1 += f"<tr style='background-color: {bg};'><td style='border: 1px solid #a4b7c6; padding: 8px;'>{row['월별']}</td><td style='border: 1px solid #a4b7c6; padding: 8px; text-align: right;'>{fmt(row['수입'])}</td><td style='border: 1px solid #a4b7c6; padding: 8px; text-align: right;'>{fmt(row['지출'])}</td><td style='border: 1px solid #a4b7c6; padding: 8px; text-align: right;'>{fmt(row['잔액'])}</td></tr>"
        st.markdown(h1 + "</table>", unsafe_allow_html=True)
        
    with tab2:
        st.subheader(f"{TARGET_YEAR}년 선교헌금 주단위결산")
        d_inc_list = [format_date_str(d) for d in df_inc_target[df_inc_target['amt'] > 0]['날짜']]
        d_exp_list = [format_date_str(d) for d in df_exp_target[df_exp_target['amt'] > 0]['날짜']]
        all_dates = sorted(list(set([d for d in d_inc_list + d_exp_list if str(d).startswith(str(TARGET_YEAR))])))
        weekly_temp = []
        cur_bal, tot_inc, tot_exp = carryover_bal, carryover_bal, 0
        for d_str in all_dates:
            inc = df_inc_target[df_inc_target['날짜'].apply(format_date_str) == d_str]['amt'].sum()
            exp = df_exp_target[df_exp_target['날짜'].apply(format_date_str) == d_str]['amt'].sum()
            cur_bal += (inc - exp); tot_inc += inc; tot_exp += exp; weekly_temp.append({"월별": d_str, "수입": inc, "지출": exp, "잔액": cur_bal})
        weekly_display = [{"월별": "합계", "수입": tot_inc, "지출": tot_exp, "잔액": tot_inc - tot_exp}] + weekly_temp[::-1] + [{"월별": "전년이월", "수입": carryover_bal, "지출": 0, "잔액": carryover_bal}]
        
        h2 = "<table style='width:100%; border-collapse: collapse; text-align: center; border: 2px solid #a4b7c6; font-size: 15px; background-color: #ffffff; color: #333333;'>"
        h2 += "<tr style='background-color: #dbe5f1;'><th style='border: 1px solid #a4b7c6; padding: 10px;'>월별</th><th style='border: 1px solid #a4b7c6; padding: 10px;'>수입</th><th style='border: 1px solid #a4b7c6; padding: 10px;'>지출</th><th style='border: 1px solid #a4b7c6; padding: 10px;'>잔액</th></tr>"
        for row in weekly_display:
            bg = "#b4c6e7" if row['월별'] == "합계" else ("#f4f5f7" if row['월별'] == "전년이월" else "#ffffff")
            h2 += f"<tr style='background-color: {bg};'><td style='border: 1px solid #a4b7c6; padding: 8px;'>{row['월별']}</td><td style='border: 1px solid #a4b7c6; padding: 8px; text-align: right;'>{fmt(row['수입'])}</td><td style='border: 1px solid #a4b7c6; padding: 8px; text-align: right;'>{fmt(row['지출'])}</td><td style='border: 1px solid #a4b7c6; padding: 8px; text-align: right;'>{fmt(row['잔액'])}</td></tr>"
        st.markdown(h2 + "</table>", unsafe_allow_html=True)

# 3. 개인별 조회
elif menu == "🔍 개인별 조회":
    names = [n for n in df_target['이름'].dropna().astype(str).str.strip().unique().tolist() if n and n != 'nan' and n != '합계']
    selected = st.selectbox("성함을 선택하세요", names)
    if selected:
        res = calculate_details(selected, df_income, df_target)
        if res:
            u_info = df_target[df_target['이름'].astype(str).str.strip() == selected]
            pos = str(u_info.iloc[0].get('직분', "")) if not u_info.empty else ""
            st.subheader(f"📄 {res['name']} ({pos})")
            st.write(f"기준일({datetime.now().strftime('%Y.%m.%d')}) / 월 작정: {int(res['commit']):,}원 / 총액: {int(res['total']):,}원")
            html = "<table style='width:100%; border-collapse: collapse; text-align: center; margin-top: 15px; background-color: #ffffff; color: #333333;'>"
            html += "<tr style='background-color: #f8f9fa;'>"
            for i in range(1, 7): html += f"<th style='border: 1px solid #ddd; padding: 10px;'>{i}월</th>"
            html += "</tr><tr>"
            for i in range(6):
                lab, amt = str(res['labs'][i]).replace('\n', '<br>'), f"{int(res['alloc'][i]):,}원" if res['alloc'][i] > 0 else "0원"
                html += f"<td style='border: 1px solid #ddd; padding: 15px;'><span style='font-size:0.85em; color:#888;'>{lab}</span><br><b>{amt}</b></td>"
            html += "</tr><tr style='background-color: #f8f9fa;'>"
            for i in range(7, 13): html += f"<th style='border: 1px solid #ddd; padding: 10px;'>{i}월</th>"
            html += "</tr><tr>"
            for i in range(6, 12):
                lab, amt = str(res['labs'][i]).replace('\n', '<br>'), f"{int(res['alloc'][i]):,}원" if res['alloc'][i] > 0 else "0원"
                html += f"<td style='border: 1px solid #ddd; padding: 15px;'><span style='font-size:0.85em; color:#888;'>{lab}</span><br><b>{amt}</b></td>"
            html += "</tr></table>"; st.markdown(html, unsafe_allow_html=True)

# 4. 인쇄용 집계표 (Admin 전용)
elif menu == "🖨️ 인쇄용 집계표":
    st.subheader("🖨️ 인쇄용 엑셀 다운로드")
    months = sorted(list(set([f"{TARGET_YEAR}{str(m).zfill(2)}" for m in range(1, 13)] + list(df_income['년월'].unique()))), reverse=True)
    target_month = st.selectbox("📌 기준월 선택", months)
    if st.button("🔄 인쇄용 파일 생성", use_container_width=True):
        with st.spinner("엑셀 파일을 생성 중입니다..."):
            donors = set(df_income[df_income['년월'] == target_month]['이름'].apply(lambda x: str(x).strip()).unique())
            with conn.session as s:
                for idx, row in df_target.iterrows():
                    name = str(row.get('이름')).strip()
                    if not name or name == 'nan' or name == '합계': continue
                    print_val = 'Y' if name in donors else 'N'
                    s.execute(text("UPDATE target SET print_yn=:p WHERE name=:n AND target_year=:y"), {"p": print_val, "n": name, "y": TARGET_YEAR})
                s.commit()
            clear_db_cache(); _, df_target_updated, _, _ = load_data(TARGET_YEAR)
            st.session_state.download_data = generate_summary_excel(df_income, df_target_updated, target_month)
            st.success(f"✅ 완성되었습니다!")
            
    if 'download_data' in st.session_state:
        st.download_button("📥 인쇄용 엑셀 다운로드", data=st.session_state.download_data, file_name=f"선교헌금_인쇄용_{target_month}.xlsx", use_container_width=True)
